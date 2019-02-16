import sys
import sip
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
import datetime
import openpyxl
import os

allplice = int(0)
paymanay = str("")
change = int(0)
analizepath = os.path.abspath('POSanalizer.xlsx')
setpath = os.path.abspath('Setting.xlsx')
jcode = int(0)
p1tag = int(0)
p2tag = int(0)
p3tag = int(0)
p4tag = int(0)
p5tag = int(0)
p6tag = int(0)
p7tag = int(0)
p8tag = int(0)
p9tag = int(0)
p10tag = int(0)
p11tag = int(0)
p12tag = int(0)
p13tag = int(0)
p14tag = int(0)
p15tag = int(0)
p16tag = int(0)
p17tag = int(0)
p18tag = int(0)
p19tag = int(0)
p20tag = int(0)
p21tag = int(0)
p22tag = int(0)
p23tag = int(0)
p24tag = int(0)
p25tag = int(0)
p26tag = int(0)
p27tag = int(0)
p28tag = int(0)
p29tag = int(0)
p30tag = int(0)
p31tag = int(0)
p32tag = int(0)
p33tag = int(0)
p34tag = int(0)
p35tag = int(0)
p36tag = int(0)

nametag1 = str("")
nametag2 = str("")
nametag3 = str("")
nametag4 = str("")
nametag5 = str("")
nametag6 = str("")
nametag7 = str("")
nametag8 = str("")
nametag9 = str("")
nametag10 = str("")
nametag11 = str("")
nametag12 = str("")
nametag13 = str("")
nametag14 = str("")
nametag15 = str("")
nametag16 = str("")
nametag17 = str("")
nametag18 = str("")
nametag19 = str("")
nametag20 = str("")
nametag21 = str("")
nametag22 = str("")
nametag23 = str("")
nametag24 = str("")
nametag25 = str("")
nametag26 = str("")
nametag27 = str("")
nametag28 = str("")
nametag29 = str("")
nametag30 = str("")
nametag31 = str("")
nametag32 = str("")
nametag33 = str("")
nametag34 = str("")
nametag35 = str("")
nametag36 = str("")

class Model(QStandardItemModel):
    def __init__(self, row, column, parent=None):
        super(Model, self).__init__(row, column, parent)
        self.row = row
        self.column = column



    def additeminrow(self, texts):
        self.insertRows(0, 1)
        for c in range(self.column):
            self.setData(self.index(0,c), texts[c])
            self.row +=1










class MainWin(QMainWindow):
    def __init__(self, parent=None):
        super(MainWin, self).__init__(parent)#GUI


        self.plicelabel = QLabel(self)
        self.plicelabel.setFrameStyle(QFrame.Box)
        self.plicelabel.setStyleSheet("font-size:50px; background-color:black; color:white;")
        self.plicelabel.setAlignment(Qt.AlignRight)
        self.plicelabel.setGeometry(800, 0, 500, 70) #label
        self.plicelabel.show()

        namelabel = QLabel('合計金額', self)
        namelabel.setStyleSheet("font-size:50px; font-family:fantasy;")
        namelabel.setGeometry(600, 0, 200, 70)

        self.paylabel = QLabel(self)
        self.paylabel.setFrameStyle(QFrame.Box)
        self.paylabel.setStyleSheet("font-size:25px; background-color:black; color:white;")
        self.paylabel.setAlignment(Qt.AlignRight)
        self.paylabel.setGeometry(800, 80, 200, 35)

        name2label = QLabel('お預金', self)
        name2label.setStyleSheet("font-size:25px; font-family:fantasy;")
        name2label.setGeometry(700, 80, 200, 35)

        self.changelabel = QLabel(self)
        self.changelabel.setFrameStyle(QFrame.Box)
        self.changelabel.setStyleSheet("font-size:25px; background-color:black; color:white;")
        self.changelabel.setAlignment(Qt.AlignRight)
        self.changelabel.setGeometry(800, 120, 200, 35)

        name3label = QLabel('お釣り', self)
        name3label.setStyleSheet("font-size:25px; font-family:fantasy;")
        name3label.setGeometry(700, 120, 200, 35)

        self.pay2label = QLabel(self)
        self.pay2label.setFrameStyle(QFrame.Box)
        self.pay2label.setStyleSheet("font-size:20px;")
        self.pay2label.setAlignment(Qt.AlignRight)
        self.pay2label.setGeometry(1075, 200, 200, 30)









        self.name1 = QPushButton('indname1', self)#//Button
        self.name1.clicked.connect(self.n1fun)
        self.name1.setGeometry(500,200,100,50)
        self.name2 = QPushButton('indname2', self)
        self.name2.clicked.connect(self.n2fun)
        self.name2.setGeometry(500,250,100,50)
        self.name3 = QPushButton('indname3', self)
        self.name3.clicked.connect(self.n3fun)
        self.name3.setGeometry(500,300,100,50)
        self.name4 = QPushButton('indname4', self)
        self.name4.clicked.connect(self.n4fun)
        self.name4.setGeometry(500,350,100,50)
        self.name5 = QPushButton('indnam5', self)
        self.name5.clicked.connect(self.n5fun)
        self.name5.setGeometry(500,400,100,50)
        self.name6 = QPushButton('indname6', self)
        self.name6.clicked.connect(self.n6fun)
        self.name6.setGeometry(500,450,100,50)
        self.name7 = QPushButton('indname7', self)
        self.name7.clicked.connect(self.n7fun)
        self.name7.setGeometry(500,500,100,50)
        self.name8 = QPushButton('indname8', self)
        self.name8.clicked.connect(self.n8fun)
        self.name8.setGeometry(500,550,100,50)
        self.name9 = QPushButton('indname9', self)
        self.name9.clicked.connect(self.n9fun)
        self.name9.setGeometry(500,600,100,50)

        self.name10 = QPushButton('indname10', self)
        self.name10.clicked.connect(self.n10fun)
        self.name10.setGeometry(600,200,100,50)
        self.name11 = QPushButton('indname11', self)
        self.name11.clicked.connect(self.n11fun)
        self.name11.setGeometry(600,250,100,50)
        self.name12 = QPushButton('indname12', self)
        self.name12.clicked.connect(self.n12fun)
        self.name12.setGeometry(600,300,100,50)
        self.name13 = QPushButton('indname13', self)
        self.name13.clicked.connect(self.n13fun)
        self.name13.setGeometry(600,350,100,50)
        self.name14 = QPushButton('indnam14', self)
        self.name14.clicked.connect(self.n14fun)
        self.name14.setGeometry(600,400,100,50)
        self.name15 = QPushButton('indname15', self)
        self.name15.clicked.connect(self.n15fun)
        self.name15.setGeometry(600,450,100,50)
        self.name16= QPushButton('indname16', self)
        self.name16.clicked.connect(self.n16fun)
        self.name16.setGeometry(600,500,100,50)
        self.name17 = QPushButton('indname17', self)
        self.name17.clicked.connect(self.n17fun)
        self.name17.setGeometry(600,550,100,50)
        self.name18 = QPushButton('indname18', self)
        self.name18.clicked.connect(self.n18fun)
        self.name18.setGeometry(600,600,100,50)

        self.name19 = QPushButton('indname1', self)#//Button
        self.name19.clicked.connect(self.n19fun)
        self.name19.setGeometry(700,200,100,50)
        self.name20 = QPushButton('indname2', self)
        self.name20.clicked.connect(self.n20fun)
        self.name20.setGeometry(700,250,100,50)
        self.name21 = QPushButton('indname3', self)
        self.name21.clicked.connect(self.n21fun)
        self.name21.setGeometry(700,300,100,50)
        self.name22 = QPushButton('indname4', self)
        self.name22.clicked.connect(self.n22fun)
        self.name22.setGeometry(700,350,100,50)
        self.name23 = QPushButton('indnam5', self)
        self.name23.clicked.connect(self.n23fun)
        self.name23.setGeometry(700,400,100,50)
        self.name24 = QPushButton('indname6', self)
        self.name24.clicked.connect(self.n24fun)
        self.name24.setGeometry(700,450,100,50)
        self.name25 = QPushButton('indname7', self)
        self.name25.clicked.connect(self.n25fun)
        self.name25.setGeometry(700,500,100,50)
        self.name26 = QPushButton('indname8', self)
        self.name26.clicked.connect(self.n26fun)
        self.name26.setGeometry(700,550,100,50)
        self.name27 = QPushButton('indname9', self)
        self.name27.clicked.connect(self.n27fun)
        self.name27.setGeometry(700,600,100,50)

        self.name28 = QPushButton('indname10', self)
        self.name28.clicked.connect(self.n28fun)
        self.name28.setGeometry(800,200,100,50)
        self.name29 = QPushButton('indname11', self)
        self.name29.clicked.connect(self.n29fun)
        self.name29.setGeometry(800,250,100,50)
        self.name30 = QPushButton('indname12', self)
        self.name30.clicked.connect(self.n30fun)
        self.name30.setGeometry(800,300,100,50)
        self.name31 = QPushButton('indname13', self)
        self.name31.clicked.connect(self.n31fun)
        self.name31.setGeometry(800,350,100,50)
        self.name32 = QPushButton('indnam14', self)
        self.name32.clicked.connect(self.n32fun)
        self.name32.setGeometry(800,400,100,50)
        self.name33 = QPushButton('indname15', self)
        self.name33.clicked.connect(self.n33fun)
        self.name33.setGeometry(800,450,100,50)
        self.name34= QPushButton('indname16', self)
        self.name34.clicked.connect(self.n34fun)
        self.name34.setGeometry(800,500,100,50)
        self.name35 = QPushButton('indname17', self)
        self.name35.clicked.connect(self.n35fun)
        self.name35.setGeometry(800,550,100,50)
        self.name36 = QPushButton('indname18', self)
        self.name36.clicked.connect(self.n36fun)
        self.name36.setGeometry(800,600,100,50)





        clear1 = QPushButton('clear', self)
        clear1.clicked.connect(self.clear1fun)
        clear1.setGeometry(500, 150, 200, 50)
        paycheckbutton = QPushButton('支払い', self)
        paycheckbutton.clicked.connect(self.payfun)
        paycheckbutton.setGeometry(900, 200, 100, 50)


        pay9 = QPushButton('9', self)
        pay9.clicked.connect(self.p9fun)
        pay9.setGeometry(1200, 250, 50, 50)
        pay8 = QPushButton('8', self)
        pay8.clicked.connect(self.p8fun)
        pay8.setGeometry(1150, 250, 50, 50)
        pay7 = QPushButton('7', self)
        pay7.clicked.connect(self.p7fun)
        pay7.setGeometry(1100, 250, 50, 50)
        pay6 = QPushButton('6', self)
        pay6.clicked.connect(self.p6fun)
        pay6.setGeometry(1200, 300, 50, 50)
        pay5 = QPushButton('5', self)
        pay5.clicked.connect(self.p5fun)
        pay5.setGeometry(1150, 300, 50, 50)
        pay4 = QPushButton('4', self)
        pay4.clicked.connect(self.p4fun)
        pay4.setGeometry(1100, 300, 50, 50)
        pay3 = QPushButton('3', self)
        pay3.clicked.connect(self.p3fun)
        pay3.setGeometry(1200, 350, 50, 50)
        pay2 = QPushButton('2', self)
        pay2.clicked.connect(self.p2fun)
        pay2.setGeometry(1150, 350, 50, 50)
        pay1 = QPushButton('1', self)
        pay1.clicked.connect(self.p1fun)
        pay1.setGeometry(1100, 350, 50, 50)
        pay00 = QPushButton('00', self)
        pay00.clicked.connect(self.p00fun)
        pay00.setGeometry(1200, 400, 50, 50)
        pay0 = QPushButton('0', self)
        pay0.clicked.connect(self.p0fun)
        pay0.setGeometry(1150, 400, 50, 50)
        pay000 = QPushButton('000', self)
        pay000.clicked.connect(self.p000fun)
        pay000.setGeometry(1100, 400, 50, 50)
        clear2 = QPushButton('clear', self)
        clear2.clicked.connect(self.clear2fun)
        clear2.setGeometry(1100, 450, 150, 50)
        setbtn = QPushButton('Setting', self)
        setbtn.clicked.connect(self.settingfun)
        setbtn.setGeometry(0, 0, 100, 50)




        self.model = Model(0, 3, self)
        self.mainlist = QTreeView(self)
        self.mainlist.setModel(self.model)
        self.mainlist.setColumnWidth(1, 140)
        self.mainlist.setGeometry(100, 100, 350, 550)













        p = self.palette()
        p.setColor(self.backgroundRole(), Qt.white)
        self.setPalette(p)
        self.setGeometry(300, 300, 250, 250)
        self.setWindowTitle('PyPOS')

#system
    def n1fun(self):
        global allplice
        global p1tag
        global nametag1
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E3"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E3"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p1tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag1)
            one = str(p1tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E3"].value
            aimpoint = analize["D3"].value
            analize["C3"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p1tag))
            analize["E3"].value = inputcode
            analize["D3"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag1)+":"+"￥"+str(p1tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()




        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')





    def n2fun(self):
        global allplice
        global p2tag
        global nametag2
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E4"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E4"].value = zaiko1
            tag.save(setpath)


            allplice = (int(allplice)+int(p2tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag2)
            one = str(p2tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E4"].value
            aimpoint = analize["D4"].value
            analize["C4"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p2tag))
            analize["E4"].value = inputcode
            analize["D4"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag2)+":"+"￥"+str(p2tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')





    def n3fun(self):
        global allplice
        global p3tag
        global nametag3
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E5"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E5"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p3tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag3)
            one = str(p3tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E5"].value
            aimpoint = analize["D5"].value
            analize["C5"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p3tag))
            analize["E5"].value = inputcode
            analize["D5"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag3)+":"+"￥"+str(p3tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n4fun(self):
        global allplice
        global p4tag
        global nametag4
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E6"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E6"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p4tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag4)
            one = str(p4tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E6"].value
            aimpoint = analize["D6"].value
            analize["C6"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p4tag))
            analize["E6"].value = inputcode
            analize["D6"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag4)+":"+"￥"+str(p4tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n5fun(self):
        global allplice
        global p5tag
        global nametag5
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E7"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E7"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p5tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag5)
            one = str(p5tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E7"].value
            aimpoint = analize["D7"].value
            analize["C7"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p5tag))
            analize["E7"].value = inputcode
            analize["D7"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag5)+":"+"￥"+str(p5tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n6fun(self):
        global allplice
        global p6tag
        global nametag6
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E8"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E8"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p6tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag6)
            one = str(p6tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E8"].value
            aimpoint = analize["D8"].value
            analize["C8"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p6tag))
            analize["E8"].value = inputcode
            analize["D8"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag6)+":"+"￥"+str(p6tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')

    def n7fun(self):
        global allplice
        global p7tag
        global nametag7
        global setpath
        global analizepath

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E9"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E9"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p7tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag7)
            one = str(p7tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E9"].value
            aimpoint = analize["D9"].value
            analize["C9"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p7tag))
            analize["E9"].value = inputcode
            analize["D9"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag7)+":"+"￥"+str(p7tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n8fun(self):
        global allplice
        global p8tag
        global nametag8
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E10"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E10"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p8tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag8)
            one = str(p8tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E10"].value
            aimpoint = analize["D10"].value
            analize["C10"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p8tag))
            analize["E10"].value = inputcode
            analize["D10"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag8)+":"+"￥"+str(p8tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()


        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')

    def n9fun(self):
        global allplice
        global p9tag
        global nametag9
        global analizepath
        global setpath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E11"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E11"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p9tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag9)
            one = str(p9tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E11"].value
            aimpoint = analize["D11"].value
            analize["C11"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p9tag))
            analize["E11"].value = inputcode
            analize["D11"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag9)+":"+"￥"+str(p9tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n10fun(self):
        global allplice
        global p10tag
        global nametag10
        global analizepath
        global setpath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E12"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E12"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p10tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag10)
            one = str(p10tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E12"].value
            aimpoint = analize["D12"].value
            analize["C12"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p10tag))
            analize["E12"].value = inputcode
            analize["D12"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag10)+":"+"￥"+str(p10tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n11fun(self):
        global allplice
        global p11tag
        global nametag11
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E13"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E13"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p11tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag11)
            one = str(p11tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E13"].value
            aimpoint = analize["D13"].value
            analize["C13"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p11tag))
            analize["E13"].value = inputcode
            analize["D13"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag11)+":"+"￥"+str(p11tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n12fun(self):
        global allplice
        global p12tag
        global nametag12
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E14"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E14"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p12tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag12)
            one = str(p12tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E14"].value
            aimpoint = analize["D14"].value
            analize["C14"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p12tag))
            analize["E14"].value = inputcode
            analize["D14"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag12)+":"+"￥"+str(p12tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n13fun(self):
        global p13tag
        global allplice
        global nametag13
        global analizepath
        global setpath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E15"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E15"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p13tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag13)
            one = str(p13tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E15"].value
            aimpoint = analize["D15"].value
            analize["C15"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p13tag))
            analize["E15"].value = inputcode
            analize["D15"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag13)+":"+"￥"+str(p13tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()


        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')

    def n14fun(self):
        global allplice
        global p14tag
        global nametag14
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E16"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E16"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p14tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag14)
            one = str(p14tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E16"].value
            aimpoint = analize["D16"].value
            analize["C16"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p14tag))
            analize["E16"].value = inputcode
            analize["D16"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag14)+":"+"￥"+str(p14tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n15fun(self):
        global allplice
        global p15tag
        global nametag15
        global analizepath
        global setpath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E17"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E17"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p15tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag15)
            one = str(p15tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E17"].value
            aimpoint = analize["D17"].value
            analize["C17"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p15tag))
            analize["E17"].value = inputcode
            analize["D17"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag15)+":"+"￥"+str(p15tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n16fun(self):
        global allplice
        global p16tag
        global nametag16
        global analizepath
        global setpath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E18"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E18"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p16tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag16)
            one = str(p16tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E18"].value
            aimpoint = analize["D18"].value
            analize["C18"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p16tag))
            analize["E18"].value = inputcode
            analize["D18"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag16)+":"+"￥"+str(p16tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')





    def n17fun(self):
        global allplice
        global p17tag
        global nametag17
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E19"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E19"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p17tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag17)
            one = str(p17tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E19"].value
            aimpoint = analize["D19"].value
            analize["C19"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p17tag))
            analize["E19"].value = inputcode
            analize["D19"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag17)+":"+"￥"+str(p17tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n18fun(self):
        global allplice
        global p18tag
        global nametag18
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E20"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E20"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p18tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag18)
            one = str(p18tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E20"].value
            aimpoint = analize["D20"].value
            analize["C20"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p18tag))
            analize["E20"].value = inputcode
            analize["D20"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag18)+":"+"￥"+str(p18tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n19fun(self):
        global allplice
        global p19tag
        global nametag19
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k3"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k3"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p19tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag19)
            one = str(p19tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E21"].value
            aimpoint = analize["D21"].value
            analize["C21"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p19tag))
            analize["E21"].value = inputcode
            analize["D21"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag19)+":"+"￥"+str(p19tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n20fun(self):
        global allplice
        global p20tag
        global nametag20
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k4"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k4"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p20tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag20)
            one = str(p20tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E22"].value
            aimpoint = analize["D22"].value
            analize["C22"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p20tag))
            analize["E22"].value = inputcode
            analize["D22"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag20)+":"+"￥"+str(p20tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n21fun(self):
        global allplice
        global p21tag
        global nametag21
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k5"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k5"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p21tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag21)
            one = str(p21tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E23"].value
            aimpoint = analize["D23"].value
            analize["C23"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p21tag))
            analize["E23"].value = inputcode
            analize["D23"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag21)+":"+"￥"+str(p21tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n22fun(self):
        global allplice
        global p22tag
        global nametag22
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k6"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k6"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p22tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag22)
            one = str(p22tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E24"].value
            aimpoint = analize["D24"].value
            analize["C24"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p22tag))
            analize["E24"].value = inputcode
            analize["D24"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag22)+":"+"￥"+str(p22tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n23fun(self):
        global allplice
        global p23tag
        global nametag23
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k7"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k7"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p23tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag23)
            one = str(p23tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E25"].value
            aimpoint = analize["D25"].value
            analize["C25"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p23tag))
            analize["E25"].value = inputcode
            analize["D25"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag23)+":"+"￥"+str(p23tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n24fun(self):
        global allplice
        global p24tag
        global nametag24
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k8"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k8"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p24tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag24)
            one = str(p24tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E26"].value
            aimpoint = analize["D26"].value
            analize["C3"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p24tag))
            analize["E26"].value = inputcode
            analize["D26"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag24)+":"+"￥"+str(p24tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n25fun(self):
        global allplice
        global p25tag
        global nametag25
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k9"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k9"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p25tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag25)
            one = str(p25tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E27"].value
            aimpoint = analize["D27"].value
            analize["C27"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p25tag))
            analize["E27"].value = inputcode
            analize["D27"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag25)+":"+"￥"+str(p25tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n26fun(self):
        global allplice
        global p26tag
        global nametag26
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k10"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k10"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p26tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag26)
            one = str(p26tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E28"].value
            aimpoint = analize["D28"].value
            analize["C28"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p26tag))
            analize["E28"].value = inputcode
            analize["D28"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag26)+":"+"￥"+str(p26tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n27fun(self):
        global allplice
        global p27tag
        global nametag27
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k11"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k11"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p27tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag27)
            one = str(p27tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E29"].value
            aimpoint = analize["D29"].value
            analize["C29"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p27tag))
            analize["E29"].value = inputcode
            analize["D29"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag27)+":"+"￥"+str(p27tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n28fun(self):
        global allplice
        global p28tag
        global nametag28
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k12"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k12"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p28tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag28)
            one = str(p28tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E30"].value
            aimpoint = analize["D30"].value
            analize["C30"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p28tag))
            analize["E30"].value = inputcode
            analize["D30"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag28)+":"+"￥"+str(p28tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n29fun(self):
        global allplice
        global p29tag
        global nametag29
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k13"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k13"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p29tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag29)
            one = str(p29tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E31"].value
            aimpoint = analize["D31"].value
            analize["C31"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p29tag))
            analize["E31"].value = inputcode
            analize["D31"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag29)+":"+"￥"+str(p29tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n30fun(self):
        global allplice
        global p30tag
        global nametag30
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k14"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k14"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p30tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag30)
            one = str(p30tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E32"].value
            aimpoint = analize["D32"].value
            analize["C32"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p30tag))
            analize["E32"].value = inputcode
            analize["D32"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag30)+":"+"￥"+str(p30tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')





    def n31fun(self):
        global allplice
        global analizepath
        global p31tag
        global nametag31
        global setpath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k15"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k15"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p31tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag31)
            one = str(p31tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E33"].value
            aimpoint = analize["D33"].value
            analize["C33"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p31tag))
            analize["E33"].value = inputcode
            analize["D33"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag31)+":"+"￥"+str(p31tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n32fun(self):
        global allplice
        global p32tag
        global nametag32
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k16"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k16"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p32tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag32)
            one = str(p32tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E34"].value
            aimpoint = analize["D34"].value
            analize["C34"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p32tag))
            analize["E34"].value = inputcode
            analize["D34"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag32)+":"+"￥"+str(p32tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n33fun(self):
        global allplice
        global p33tag
        global nametag33
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k17"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k17"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p33tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag33)
            one = str(p33tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E35"].value
            aimpoint = analize["D35"].value
            analize["C35"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p33tag))
            analize["E35"].value = inputcode
            analize["D35"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag33)+":"+"￥"+str(p33tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n34fun(self):
        global allplice
        global p34tag
        global nametag34
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k18"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k18"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p34tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag34)
            one = str(p34tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E36"].value
            aimpoint = analize["D36"].value
            analize["C36"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p34tag))
            analize["E36"].value = inputcode
            analize["D36"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag34)+":"+"￥"+str(p34tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n35fun(self):
        global allplice
        global p35tag
        global nametag35
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k19"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k19"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p35tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag35)
            one = str(p35tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E37"].value
            aimpoint = analize["D37"].value
            analize["C37"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p35tag))
            analize["E37"].value = inputcode
            analize["D37"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag35)+":"+"￥"+str(p35tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n36fun(self):
        global allplice
        global p36tag
        global nametag36
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k20"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k20"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p36tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag36)
            one = str(p36tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E38"].value
            aimpoint = analize["D38"].value
            analize["C38"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p36tag))
            analize["E38"].value = inputcode
            analize["D38"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag36)+":"+"￥"+str(p36tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')

    def payfun(self):
        global paymanay
        global allplice
        global change
        change = (int(paymanay)-int(allplice))
        self.changelabel.setText(str(change))
        t = datetime.datetime.now()
        time = t.time()
        string = str('=====総額=======')
        one = str(allplice)
        display_list = [str(time), string, one]
        self.model.additeminrow(display_list)

        datatime = datetime.datetime.now()
        datasorce = (str(datatime)+":===総額===:"+"￥"+str(allplice)+"\n")
        data = os.path.abspath("取引データ.txt")
        file = open(data, 'a')
        file.write(datasorce)
        file.close()



#paysystem

    def p9fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('9'))
        self.paylabel.setText(str(paymanay))
        self.pay2label.setText(str(paymanay))

    def p8fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('8'))
        self.paylabel.setText(str(paymanay))
        self.pay2label.setText(str(paymanay))

    def p7fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('7'))
        self.paylabel.setText(str(paymanay))
        self.pay2label.setText(str(paymanay))

    def p6fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('6'))
        self.paylabel.setText(str(paymanay))
        self.pay2label.setText(str(paymanay))

    def p5fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('5'))
        self.paylabel.setText(str(paymanay))
        self.pay2label.setText(str(paymanay))

    def p4fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('4'))
        self.paylabel.setText(str(paymanay))
        self.pay2label.setText(str(paymanay))

    def p3fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('3'))
        self.paylabel.setText(str(paymanay))
        self.pay2label.setText(str(paymanay))

    def p2fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('2'))
        self.paylabel.setText(str(paymanay))
        self.pay2label.setText(str(paymanay))

    def p1fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('1'))
        self.paylabel.setText(str(paymanay))
        self.pay2label.setText(str(paymanay))

    def p00fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('00'))
        self.paylabel.setText(str(paymanay))
        self.pay2label.setText(str(paymanay))

    def p0fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('0'))
        self.paylabel.setText(str(paymanay))
        self.pay2label.setText(str(paymanay))

    def p000fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('000'))
        self.paylabel.setText(str(paymanay))
        self.pay2label.setText(str(paymanay))

    def clear2fun(self):
        global paymanay
        paymanay = (str(''))
        self.paylabel.setText(str(paymanay))
        self.pay2label.setText(str(paymanay))
        t = datetime.datetime.now()

    def clear1fun(self):
        global allplice
        allplice = (int(0))
        self.plicelabel.setText(str(allplice))
        t = datetime.datetime.now()
        time = t.time()
        string = str("クリア")
        one = str("")
        display_list = [str(time), string, one]
        self.model.additeminrow(display_list)

        datatime = datetime.datetime.now()
        datasorce = (str(datatime)+":以下クリア:"+"￥0"+"\n")
        data = os.path.abspath("取引データ.txt")
        file = open(data, 'a')
        file.write(datasorce)
        file.close()

#Key
    def keyPressEvent(self, e):
        global paymanay
        global allplice
        global change

        if e.key() == Qt.Key_1:
            paymanay = (str(paymanay)+str('1'))
            self.paylabel.setText(str(paymanay))
            self.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_2:
            paymanay = (str(paymanay)+str('2'))
            self.paylabel.setText(str(paymanay))
            self.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_3:
            paymanay = (str(paymanay)+str('3'))
            self.paylabel.setText(str(paymanay))
            self.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_4:
            paymanay = (str(paymanay)+str('4'))
            self.paylabel.setText(str(paymanay))
            self.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_5:
            paymanay = (str(paymanay)+str('5'))
            self.paylabel.setText(str(paymanay))
            self.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_6:
            paymanay = (str(paymanay)+str('6'))
            self.paylabel.setText(str(paymanay))
            self.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_7:
            paymanay = (str(paymanay)+str('7'))
            self.paylabel.setText(str(paymanay))
            self.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_8:
            paymanay = (str(paymanay)+str('8'))
            self.paylabel.setText(str(paymanay))
            self.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_9:
            paymanay = (str(paymanay)+str('9'))
            self.paylabel.setText(str(paymanay))
            self.pay2label.setText(str(paymanay))


        if e.key() == Qt.Key_0:
            paymanay = (str(paymanay)+str('0'))
            self.paylabel.setText(str(paymanay))
            self.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_Control:
            allplice = (int(0))
            paymanay = (str(''))
            change = (int(0))
            self.paylabel.setText(str(paymanay))
            self.pay2label.setText(str(paymanay))
            self.plicelabel.setText(str(allplice))
            self.changelabel.setText(str(change))
            t = datetime.datetime.now()
            time = t.time()
            string = str("クリア")
            one = str("")
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":以下クリア:"+"￥0"+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()



        if e.key() == Qt.Key_Return:
            change = (int(paymanay)-int(allplice))
            self.changelabel.setText(str(change))
            t = datetime.datetime.now()
            time = t.time()
            string = str('=====総額=======')
            one = str(allplice)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":===総額===:"+"￥"+str(allplice)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()


    def settingfun(self):

        global nametag1
        global nametag2
        global nametag3
        global nametag4
        global nametag5
        global nametag6
        global nametag7
        global nametag8
        global nametag9
        global nametag10
        global nametag11
        global nametag12
        global nametag13
        global nametag14
        global nametag15
        global nametag16
        global nametag17
        global nametag18
        global nametag19
        global nametag20
        global nametag21
        global nametag22
        global nametag23
        global nametag24
        global nametag25
        global nametag26
        global nametag27
        global nametag28
        global nametag29
        global nametag30
        global nametag31
        global nametag32
        global nametag33
        global nametag34
        global nametag35
        global nametag36




        global setpath
        global p1tag
        global p2tag
        global p3tag
        global p4tag
        global p5tag
        global p7tag
        global p8tag
        global p9tag
        global p10tag
        global p11tag
        global p12tag
        global p13tag
        global p14tag
        global p15tag
        global p16tag
        global p17tag
        global p18tag
        global p18tag
        global p19tag
        global p20tag
        global p21tag
        global p22tag
        global p23tag
        global p24tag
        global p25tag
        global p26tag
        global p27tag
        global p28tag
        global p29tag
        global p30tag
        global p31tag
        global p32tag
        global p33tag
        global p34tag
        global p35tag
        global p36tag
        tag = openpyxl.load_workbook(setpath)
        act = tag.active
        #ここから商品名

        nametag1 = act["B3"].value
        nametag2 = act["B4"].value
        nametag3 = act["B5"].value
        nametag4 = act["B6"].value
        nametag5 = act["B7"].value
        nametag6 = act["B8"].value
        nametag7 = act["B9"].value
        nametag8 = act["B10"].value
        nametag9 = act["B11"].value
        nametag10 = act["B12"].value
        nametag11 = act["B13"].value
        nametag12 = act["B14"].value
        nametag13 = act["B15"].value
        nametag14 = act["B16"].value
        nametag15 = act["B17"].value
        nametag16 = act["B18"].value
        nametag17 = act["B19"].value
        nametag18 = act["B20"].value


        nametag19 = act["H3"].value
        nametag20 = act["H4"].value
        nametag21 = act["H5"].value
        nametag22 = act["H6"].value
        nametag23 = act["H7"].value
        nametag24 = act["H8"].value
        nametag25 = act["H9"].value
        nametag26 = act["H10"].value
        nametag27 = act["H11"].value
        nametag28 = act["H12"].value
        nametag29 = act["H13"].value
        nametag30 = act["H14"].value
        nametag31 = act["H15"].value
        nametag32 = act["H16"].value
        nametag33 = act["H17"].value
        nametag34 = act["H18"].value
        nametag35 = act["H19"].value
        nametag36 = act["H20"].value


        p1tag = act["C3"].value
        p2tag = act["C4"].value
        p3tag = act["C5"].value
        p4tag = act["C6"].value
        p5tag = act["C7"].value
        p6tag = act["C8"].value
        p7tag = act["C9"].value
        p8tag = act["C10"].value
        p9tag = act["C11"].value
        p10tag = act["C12"].value
        p11tag = act["C13"].value
        p12tag = act["C14"].value
        p13tag = act["C15"].value
        p14tag = act["C16"].value
        p15tag = act["C17"].value
        p16tag = act["C18"].value
        p17tag = act["C19"].value
        p18tag = act["C20"].value

        p19tag = act["I3"].value
        p20tag = act["I4"].value
        p21tag = act["I5"].value
        p22tag = act["I6"].value
        p23tag = act["I7"].value
        p24tag = act["I8"].value
        p25tag = act["I9"].value
        p26tag = act["I10"].value
        p27tag = act["I11"].value
        p28tag = act["I12"].value
        p29tag = act["I13"].value
        p30tag = act["I14"].value
        p31tag = act["I15"].value
        p32tag = act["I16"].value
        p33tag = act["I17"].value
        p34tag = act["I18"].value
        p35tag = act["I19"].value
        p36tag = act["I20"].value

        self.name1.setText(str(nametag1))
        self.name2.setText(str(nametag2))
        self.name3.setText(str(nametag3))
        self.name4.setText(str(nametag4))
        self.name5.setText(str(nametag5))
        self.name6.setText(str(nametag6))
        self.name7.setText(str(nametag7))
        self.name8.setText(str(nametag8))
        self.name9.setText(str(nametag9))
        self.name10.setText(str(nametag10))
        self.name11.setText(str(nametag11))
        self.name12.setText(str(nametag12))
        self.name13.setText(str(nametag13))
        self.name14.setText(str(nametag14))
        self.name15.setText(str(nametag15))
        self.name16.setText(str(nametag16))
        self.name17.setText(str(nametag17))
        self.name18.setText(str(nametag18))

        self.name19.setText(str(nametag19))
        self.name20.setText(str(nametag20))
        self.name21.setText(str(nametag21))
        self.name22.setText(str(nametag22))
        self.name23.setText(str(nametag23))
        self.name24.setText(str(nametag24))
        self.name25.setText(str(nametag25))
        self.name26.setText(str(nametag26))
        self.name27.setText(str(nametag27))
        self.name28.setText(str(nametag28))
        self.name29.setText(str(nametag29))
        self.name30.setText(str(nametag30))
        self.name31.setText(str(nametag31))
        self.name32.setText(str(nametag32))
        self.name33.setText(str(nametag33))
        self.name34.setText(str(nametag34))
        self.name35.setText(str(nametag35))
        self.name36.setText(str(nametag36))







































app = QApplication(sys.argv)
main_window = MainWin()
main_window.show()
main_window.raise_()
sys.exit(app.exec_())
main()
